import io
import re
import logging
import openpyxl

logger = logging.getLogger(__name__)

_TRUST_KEYWORDS = ("trust", "entity", "name")
_ACCT_KEYWORDS  = ("account", "fund", "acct", "#")


class ExcelIndex:
    def __init__(self) -> None:
        self._index: dict[str, str] = {}

    @classmethod
    def from_bytes_list(cls, files: list[tuple[str, bytes]]) -> "ExcelIndex":
        idx = cls()
        for name, data in files:
            try:
                wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
                idx._load_workbook(wb, name)
            except Exception as exc:
                logger.warning("Could not parse '%s': %s", name, exc)
        logger.info("Excel index built from %d Box checklists: %d account mappings", len(files), len(idx))
        return idx

    def _load_workbook(self, wb, source: str) -> None:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(c).lower().strip() if c else "" for c in rows[0]]
            trust_col = self._col(headers, _TRUST_KEYWORDS)
            acct_col  = self._col(headers, _ACCT_KEYWORDS)
            if trust_col is None or acct_col is None:
                continue
            for row in rows[1:]:
                trust = row[trust_col] if len(row) > trust_col else None
                acct  = row[acct_col]  if len(row) > acct_col  else None
                if not trust or not acct:
                    continue
                key = self._normalize(str(acct))
                if key and len(key) >= 4:
                    self._index[key] = str(trust).strip()

    @staticmethod
    def _col(headers: list[str], keywords: tuple) -> int | None:
        for i, h in enumerate(headers):
            if any(k in h for k in keywords):
                return i
        return None

    @staticmethod
    def _normalize(account: str) -> str:
        return re.sub(r"[^A-Z0-9]", "", account.upper())

    def lookup_trust(self, account_number: str) -> str | None:
        key = self._normalize(account_number)
        if key in self._index:
            return self._index[key]
        if len(key) < 6:
            return None
        matches = {
            trust
            for k, trust in self._index.items()
            if len(k) >= 6 and (k in key or key in k)
        }
        if len(matches) == 1:
            return matches.pop()
        if len(matches) > 1:
            logger.warning(
                "Account '%s' substring-matches %d different trusts in Excel — ambiguous, skipping.",
                account_number, len(matches),
            )
        return None

    def __len__(self) -> int:
        return len(self._index)
